#!/usr/bin/env python3
"""
Project Breakdowns — build_summary.py
Generates a presale pricing summary Excel workbook from structured JSON data.

Usage:
    python build_summary.py --data project_data.json --output "Harth_Pricing_Summary.xlsx"

Methodology (MLA standard, per Melissa Nestoruk):
  - Per row: Avg Interior SF = AVERAGE(min SF, max SF); Avg Price = AVERAGE(starting, top-end)
  - Weighted Avg SF    = SUMPRODUCT(avg SFs, counts) / total unit count
  - Weighted Avg Price = SUMPRODUCT(avg prices, counts) / total unit count
  - Average PSF        = Weighted Avg Price / Weighted Avg SF
All computed as live Excel formulas so edits to counts/prices/sizes recalculate.
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────────────────────

YELLOW     = PatternFill("solid", fgColor="FFFF00")
LIGHT_GREY = PatternFill("solid", fgColor="D9D9D9")
HEADER_BG  = PatternFill("solid", fgColor="404040")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
NORMAL_FONT  = Font(name="Calibri", size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=13)

THIN_SIDE  = Side(style="thin")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

MONEY_FMT = '"$"#,##0'
PSF_FMT   = '"$"#,##0.00'
SF_FMT    = '#,##0.0'


def relabel(header: str, price_label: str) -> str:
    """Swap the generic 'Price'/'PPSF' wording for a project-specific label,
    e.g. price_label='Rent (Monthly)' turns 'Starting Price' into
    'Starting Rent (Monthly)' and 'Starting PPSF' into 'Starting Rent PSF'."""
    if price_label == "Price":
        return header
    return header.replace("PPSF", "Rent PSF").replace("Price", price_label)


def apply_header(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_BG
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def apply_data(cell, value, align=LEFT, number_format=None, fill=None, bold=False):
    cell.value = value
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.alignment = align
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill


# ── Summary Sheet ────────────────────────────────────────────────────────────

SUMMARY_HEADERS = [
    "Unit Type",            # A
    "Plan Name(s)",         # B
    "Count",                # C
    "Min Interior SF",      # D
    "Max Interior SF",      # E
    "Avg Interior SF",      # F
    "Starting Price",       # G
    "Est. Top-End Price",   # H
    "Avg Price",            # I
    "Starting PPSF",        # J
]

COL_WIDTHS_SUMMARY = [18, 26, 9, 15, 15, 15, 17, 18, 17, 14]


def build_summary_sheet(ws, project_name: str, rows: list[dict], price_label: str = "Price"):
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = project_name
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    title_cell.fill = LIGHT_GREY

    # Subtitle
    ws.merge_cells("A2:J2")
    ws["A2"].value = "Pricing Summary — Unit Type Breakdown"
    ws["A2"].font = BOLD_FONT
    ws["A2"].alignment = CENTER

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for i, w in enumerate(COL_WIDTHS_SUMMARY, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row (row 4)
    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        apply_header(ws.cell(row=4, column=col_idx), relabel(header, price_label))
    ws.row_dimensions[4].height = 20

    # Data rows (starting row 5)
    data_start = 5
    for row_num, r in enumerate(rows, start=data_start):
        apply_data(ws.cell(row=row_num, column=1), r.get("unit_type", ""))
        apply_data(ws.cell(row=row_num, column=2), r.get("plan_names", ""))

        # Count cell — yellow (editable input)
        c_count = ws.cell(row=row_num, column=3)
        c_count.value = r.get("count") or 0
        c_count.font = NORMAL_FONT
        c_count.fill = YELLOW
        c_count.alignment = CENTER
        c_count.border = THIN_BORDER

        apply_data(ws.cell(row=row_num, column=4), r.get("min_sf"), align=CENTER)
        apply_data(ws.cell(row=row_num, column=5), r.get("max_sf"), align=CENTER)

        # Avg Interior SF = mid of min/max (live formula)
        apply_data(ws.cell(row=row_num, column=6),
                   f"=AVERAGE(D{row_num}:E{row_num})", align=CENTER, number_format=SF_FMT)

        apply_data(ws.cell(row=row_num, column=7), r.get("starting_price"),
                   align=RIGHT, number_format=MONEY_FMT)
        apply_data(ws.cell(row=row_num, column=8), r.get("top_end_price"),
                   align=RIGHT, number_format=MONEY_FMT)

        # Avg Price = mid of starting/top-end (AVERAGE ignores a blank top-end)
        apply_data(ws.cell(row=row_num, column=9),
                   f"=AVERAGE(G{row_num}:H{row_num})", align=RIGHT, number_format=MONEY_FMT)

        # Starting PPSF = Starting Price / Min SF
        apply_data(ws.cell(row=row_num, column=10),
                   f"=IF(D{row_num}>0,G{row_num}/D{row_num},\"-\")",
                   align=RIGHT, number_format=PSF_FMT)

        ws.row_dimensions[row_num].height = 18

    data_end = data_start + len(rows) - 1
    c_rng = f"C{data_start}:C{data_end}"
    f_rng = f"F{data_start}:F{data_end}"
    i_rng = f"I{data_start}:I{data_end}"

    # ── Weighted Averages block (MLA standard methodology) ──
    block_header = data_end + 2
    ws.merge_cells(f"A{block_header}:J{block_header}")
    ws[f"A{block_header}"].value = "Weighted Averages — SUMPRODUCT(average × unit count) / total unit count"
    ws[f"A{block_header}"].font = BOLD_FONT

    sf_row = block_header + 2     # Weighted Avg Interior SF row
    price_row = block_header + 3  # Weighted Avg Price row
    block_rows = [
        ("Total Units", f"=SUM({c_rng})", "#,##0"),
        ("Weighted Avg Interior SF", f"=SUMPRODUCT({f_rng},{c_rng})/SUM({c_rng})", SF_FMT),
        (relabel("Weighted Avg Price (Mid of Starting/Top-End)", price_label),
         f"=SUMPRODUCT({i_rng},{c_rng})/SUM({c_rng})", MONEY_FMT),
        (relabel("Average PSF (Weighted Avg Price / Weighted Avg SF)", price_label),
         f"=J{price_row}/J{sf_row}", PSF_FMT),
    ]

    for i, (label, formula, fmt) in enumerate(block_rows):
        r = block_header + 1 + i
        ws.merge_cells(f"A{r}:I{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font = NORMAL_FONT
        c.alignment = RIGHT
        for col in range(1, 11):
            ws.cell(row=r, column=col).fill = LIGHT_GREY
            ws.cell(row=r, column=col).border = THIN_BORDER
        v = ws.cell(row=r, column=10)
        v.value = formula
        v.number_format = fmt
        v.font = BOLD_FONT
        v.alignment = RIGHT
        ws.row_dimensions[r].height = 18

    # Yellow note below
    note_row = block_header + len(block_rows) + 2
    ws.merge_cells(f"A{note_row}:J{note_row}")
    ws[f"A{note_row}"].value = ("* Yellow cells (Count) are editable — all averages recalculate automatically. "
                                "Avg SF = mid of min/max; Avg Price = mid of starting/top-end.")
    ws[f"A{note_row}"].font = Font(name="Calibri", size=9, italic=True, color="666666")
    ws[f"A{note_row}"].fill = PatternFill("solid", fgColor="FFFACD")


# ── Plan Detail Sheet ────────────────────────────────────────────────────────

DETAIL_HEADERS = [
    "Unit Type",
    "Plan Name",
    "Count",
    "Interior SF",
    "Starting Price",
    "Starting PPSF",
]

COL_WIDTHS_DETAIL = [18, 16, 9, 14, 18, 14]


def build_detail_sheet(ws, project_name: str, rows: list[dict], price_label: str = "Price"):
    ws.title = "Plan Detail"

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{project_name} — Plan Detail"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    title_cell.fill = LIGHT_GREY
    ws.row_dimensions[1].height = 24

    for i, w in enumerate(COL_WIDTHS_DETAIL, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, header in enumerate(DETAIL_HEADERS, start=1):
        apply_header(ws.cell(row=2, column=col_idx), relabel(header, price_label))
    ws.row_dimensions[2].height = 20

    detail_row = 3
    for r in rows:
        # Expand plan_names into individual rows if comma-separated
        plan_names_raw = r.get("plan_names", "")
        plans = [p.strip() for p in plan_names_raw.split(",")] if plan_names_raw else [""]

        min_sf = r.get("min_sf")
        max_sf = r.get("max_sf")
        start_price = r.get("starting_price")
        count = r.get("count", 0)
        unit_type = r.get("unit_type", "")

        # For multi-plan types, distribute count evenly (rough split)
        per_plan_count = round(count / len(plans)) if plans else count

        for plan in plans:
            apply_data(ws.cell(row=detail_row, column=1), unit_type)
            apply_data(ws.cell(row=detail_row, column=2), plan)
            c_count = ws.cell(row=detail_row, column=3)
            c_count.value = per_plan_count
            c_count.font = NORMAL_FONT
            c_count.fill = YELLOW
            c_count.alignment = CENTER
            c_count.border = THIN_BORDER

            # SF — show range if min != max, else single value
            if min_sf and max_sf and min_sf != max_sf:
                sf_display = f"{min_sf}–{max_sf}"
            else:
                sf_display = min_sf or ""
            apply_data(ws.cell(row=detail_row, column=4), sf_display, align=CENTER)

            apply_data(ws.cell(row=detail_row, column=5), start_price,
                       align=RIGHT, number_format=MONEY_FMT)

            psf = round(start_price / min_sf, 2) if (min_sf and start_price) else ""
            apply_data(ws.cell(row=detail_row, column=6), psf,
                       align=RIGHT, number_format=PSF_FMT)

            ws.row_dimensions[detail_row].height = 18
            detail_row += 1


# ── Main ─────────────────────────────────────────────────────────────────────

def build(data_path: str, output_path: str):
    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    project_name = data.get("project_name", "Project")
    rows = data.get("rows", [])
    # Set "price_label" in the JSON to relabel columns for rental projects,
    # e.g. "Rent (Monthly)" — prices then read as monthly rents throughout.
    price_label = data.get("price_label", "Price")

    wb = Workbook()
    ws_summary = wb.active
    build_summary_sheet(ws_summary, project_name, rows, price_label)

    ws_detail = wb.create_sheet()
    build_detail_sheet(ws_detail, project_name, rows, price_label)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build a presale pricing summary Excel workbook.")
    parser.add_argument("--data",   required=True, help="Path to JSON data file")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    args = parser.parse_args()
    build(args.data, args.output)
